import openpyxl
import json
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw

wb = openpyxl.load_workbook('math.xlsx')

sheet = wb["Form Responses 1"]

with open("cell color.json", "r") as file:
    cell_color = json.load(file)

#print(cell_color.fill.fgColor.index, cnt_row, cnt_column)

data = list()
with open("emails.json", "w") as file:
    json.dump([], file)
with open("emails.json", "r") as file:
    emails = json.load(file)

for row in range(2,51):
    cell_obj=sheet.cell(column=2,row=row)
    if cell_obj.fill.fgColor.index in cell_color:
        #print(sheet.cell(column=2,row=row).value)
        email = sheet.cell(column=8, row=row).value,
        if email in emails:
            continue
        emails += email,
        data+=[],

        for col in [2,8,6,12,11]:
            co=sheet.cell(column=col,row=row)
            co.value=co.value.strip()
            if col==12:
                data[-1]+=("Nagad" if co.value[-2]=='d' else "Bkash"),
                continue
            if col==6:
                data[-1]+=co.value.replace("+88","").replace(" ","").replace("-",""),
                continue
            if col==8 and co.value[-4:]!=".com":
                co.value+=".com"
            data[-1]+=co.value,

with open("emails.json", "w") as file:
    json.dump(emails, file)

oimg =  Image.open("poster.png")
font = ImageFont.truetype("font/DroidSansMono.ttf", 30)
Yaxis = [670 , 807, 946, 1084, 1224]
Xaxis = 747
for el in data:
    img= oimg.copy()
    draw = ImageDraw.Draw(img)
    name = el[0]
    for y in range(5):
        val = el[y]
        draw.text((Xaxis, Yaxis[y]), " " * ((30 - len(val)) // 2) + val, (0, 0, 0), font=font)
    img.save(f"Generated Poster/{name}.png")

"""with open("reg data.json", "w") as file:
    json.dump(data, file)
    
with open("reg data.json","r") as file:
    print(json.load(file))"""